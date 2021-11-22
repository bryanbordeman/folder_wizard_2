from openpyxl import load_workbook
import sys
import time

def main():
    # print('working')
    inputs = sys.argv[1] # input string
    # inputDict = eval('dict('+inputs+')') # convert input string into dict
    # create_project_log(inputDict)
    print(inputs)

def create_project_log(inputs):
        '''update quote log'''
        log = inputs['log']
        book = load_workbook(log)
        ws = book.worksheets[0]
        
        # get fist empty cell
        for cell in ws["B"]:
            if cell.value is None:
                current_row = cell.row
                break
        else:
            cell.row + 1

        # place inputs into spreadsheet
        ws[f'A{current_row}'] = f"{inputs['p_project_number']}{inputs['p_project_billing']}"
        ws[f'B{current_row}'] = f"{inputs['p_project_state']}"
        ws[f'C{current_row}'] = f"{inputs['p_project_order']}"
        ws[f'D{current_row}'] = f"{inputs['p_project_name']} {inputs['p_project_type']}"
        ws[f'E{current_row}'] = f"{inputs['p_labor']}"
        ws[f'F{current_row}'] = time.strftime("%D")
        ws[f'G{current_row}'] = inputs['p_project_type']

        book.save(log)
        book.close()
    
if __name__ == "__main__":
    main()