'''==========================================
Title:  folder_wizard.py
Author:  Bryan Bordeman
Start Date:  062219
Updated:  041521
Version:  v3.3.2

;=========================================='''


import os
import time
from datetime import datetime
from datetime import timedelta
from get_next_num import get_next_quote_num
from get_next_num import get_next_project_num
from get_next_num import get_next_service_num
from openpyxl import*
from openpyxl.styles import PatternFill, Font, Color
from zip_2_state import find_state
from project_attribute_list import*
import sqlite3
import subprocess

# global var ----------------------------
current_year = time.strftime("%Y")
database = 'protaskinate.db'

# below path is for testing at home
# project_dir = r'C:\Users\Bryan\Google Drive\Programming\Python\folder_wizard\Global'
# opportunity_dir = r"C:\Users\Bryan\Google Drive\Programming\Python\folder_wizard\RFQ's"

# below path is live on office server
project_dir = r'T:\Global'
opportunity_dir = r"\\servergps\Documents on Server\RFQ's"

def main():
    pass
# below is for testing only
    # quote_data = 'opportunity.pkl'
    # quote_obj = {}  # if quote_data does not exist
    # if os.path.exists(quote_data):
    #     with open(quote_data, 'rb') as rfp:
    #         quote_obj = pickle.load(rfp)
    # print(quote_obj)


class Opportunity(object):
    def __init__(self, project_name, project_category, project_type, type_code, project_zip, customer_list, bid_due, manager):
        self.quote_number = get_next_quote_num()
        self.project_name = project_name
        self.project_category = project_category
        self.project_type = project_type
        self.type_code = type_code
        self.project_zip = project_zip
        self.customer_list = customer_list
        self.bid_due = bid_due
        self.manager = manager


        quote = f'{self.quote_number} {self.manager} {self.project_name} {self.type_code}'

        self.write_log()
        self.create_table()
        self.write_opp_data()
        # self.pickle_opp()

        # make new year folder if current year does not match dir list-------------
        opportunity_dir_list = (os.listdir(opportunity_dir))
        print(opportunity_dir_list)
        year_list = []
        

        for year in opportunity_dir_list:
            try:
                if int(year[:5]):
                    year_list.append(year)
            except ValueError:
                continue
        year_list.sort()
        if year_list[-1] == str(current_year):
            year_dir = current_year
        else:
            self.createFolder(f'{opportunity_dir}/{current_year} Quotes')

        #-----------------------------------------------------------------------

        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/00_quotations_estimates')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/00_quotations_estimates/vendor_quotes')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/01_drawings_specs')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/02_rfi_addenda')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/03_photos')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/04_misc_docs')
        
        windows_user = str(os.getlogin())

        if windows_user != 'mark':
            self.make_readme(quote)


        task_list = [['Proposal / Estimate', 'Bryan Bordeman', self.bid_due],
                     ['Follow-up on Quote', 'Bryan Bordeman', (datetime.strptime(self.bid_due, '%m/%d/%y').date() + timedelta(days=7)).strftime("%#m/%#d/%y")]]

                    #  ['Review Proposal', 'Mark Holder', (datetime.strptime(self.bid_due, '%m/%d/%y').date() - timedelta(days=2)).strftime("%#m/%#d/%y")]

        self.create_task(self.quote_number, task_list)
        self.path = f"{opportunity_dir}\{current_year} Quotes\{quote}"

    def __repr__(self):
        return f'{self.quote_number} {self.manager} {self.project_name} {self.type_code}'

    def open_folder(self):
        subprocess.Popen(fr"explorer {self.path}")

    def create_table(self):
        '''Creates table from database if none exist'''
        self.execute_db_query(f"""CREATE TABLE IF NOT EXISTS 'opportunity' (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    quote_number text,
                    project_name text,
                    project_category text,
                    project_type text,
                    type_code text,
                    project_zip text,
                    customer_list text,
                    bid_due text
                    );""")

        self.execute_db_query(f"""CREATE TABLE IF NOT EXISTS 'project' (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    project_number text,
                    project_name text,
                    project_category text,
                    project_type text,
                    type_code text,
                    project_zip text,
                    customer text,
                    quote text,
                    terms text,
                    tax text,
                    billing text,
                    labor_code text,
                    order_type text,
                    price text
                    );""")

    def execute_db_query(self, query, parameters=()):
        with sqlite3.connect(database) as conn:
            cursor = conn.cursor()
            query_result = cursor.execute(query, parameters)
            conn.commit()
        return query_result



    def write_opp_data(self):
        '''write opportunity to database'''

        table = 'opportunity'

        self.execute_db_query(f"""INSERT INTO {table} (
                                    quote_number,
                                    project_name,
                                    project_category,
                                    project_type,
                                    type_code,
                                    project_zip,
                                    customer_list,
                                    bid_due
                                    ) VALUES(?, ?, ?, ?, ?, ?, ?, ?);""", (str(self.quote_number), str(self.project_name), str(self.project_category), str(self.project_type), str(self.type_code), str(self.project_zip), str(self.customer_list), str(self.bid_due)))

    def fetch_opp_data(self, quote):
        '''fetch opportunity from database'''

        table = 'opportunity'

        table_str = self.execute_db_query(
            f"""SELECT * FROM {table} WHERE quote_number = '{quote}';""")  # makes

        try:
            opp_list = list(table_str)[0]  # make table data into list
        except IndexError:
            opp_list = []

        return opp_list

        # quote_number = opp_list[1]
        # project_name = opp_list[2]
        # project_category = opp_list[3]
        # project_type = opp_list[4]
        # type_code = opp_list[5]
        # project_zip = opp_list[6]
        # bid_due = opp_list[8]

    def write_log(self):
        '''update quote log'''
        log = "Master Quote Log"
        completeName = os.path.join(opportunity_dir, log + ".xlsx")
        book = load_workbook(completeName)
        ws = book.worksheets[0]
        for cell in ws["A"]:
            if cell.value is None:
                print(cell.row)
                current_row = cell.row
                break
        else:
            cell.row + 1

        ws[f'A{current_row}'] = self.quote_number
        ws[f'B{current_row}'] = self.manager
        ws[f'C{current_row}'] = time.strftime("%D")
        ws[f'D{current_row}'] = f'{self.project_name} {self.type_code}'
        ws[f'E{current_row}'] = find_state(int(self.project_zip))
        ws[f'F{current_row}'] = self.bid_due

        #NOTE below is code to fill cell with green background

        # for col in ['A','B','C','D','E','F','G','H']:
        #     ws[f'{col}{current_row}'].fill = PatternFill(
        #         start_color="92d050", end_color="92d050", fill_type="solid")

        book.save(completeName)
        book.close()

    def createFolder(self, directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print('Error: Creating directory. ' + directory)

    def make_readme(self, quote):
        name_of_file = 'README'
        path = f'{opportunity_dir}/{current_year} Quotes/{quote}'
        completeName = os.path.join(path, name_of_file + ".txt")

        readme = open(completeName, "w")

        opportunity_info = f'Quote Number = {self.quote_number}\nProject Name = {self.project_name}\nProject Category = {self.project_category}\nProject Type = {self.project_type}\nProject Zip = {self.project_zip}\nBid Due Date = {self.bid_due}\nCustomer List = {self.customer_list}'

        # f'Project Number = {self.project_number}\nProject Name = {self.project_name}\nProject Category = {self.project_category}\nProject Type = {self.project_type}\nProject Zip = {self.project_zip}\nCustomer = {self.customer}\nQuote Number = {self.quote}\nTerms = {self.terms}\nTax Exempt = {self.tax}\nBilling Type = {self.billing}\nSell Price (USD) = ${self.price}\n'
        readme.write(opportunity_info)
        readme.close()

    def create_task(self, number, task_list):
        table = 'tasks'

        self.col_list = {'Project Number': "number",
                         'Task Tilte': "title",
                         'Assigned To': "assigned_to",
                         'Status': "status",
                         '% Complete': "complete",
                         'Start Date': "start_date",
                         'Due Date': "due_date",
                         'Description': "description",
                         'Date Completed': "date_completed",
                         'Username': "username"}

        #static var _____________________________________________________
        username = 'Mr. Robot'
        description = 'Automatically generated from Folder Wizard'
        start = time.strftime("%#m/%#d/%y")
        status = 'Not Started'
        complete = '0%'

        for i in range(len(task_list)):
            self.execute_db_query(f"""INSERT INTO {table}
            ({list(self.col_list.values())[0]},
            {list(self.col_list.values())[1]},
            {list(self.col_list.values())[2]},
            {list(self.col_list.values())[3]},
            {list(self.col_list.values())[4]},
            {list(self.col_list.values())[5]},
            {list(self.col_list.values())[6]},
            {list(self.col_list.values())[7]},
            {list(self.col_list.values())[9]}) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?);""", (number, task_list[i][0], task_list[i][1], status, complete, start, task_list[i][2], description, username))

class Project(Opportunity):
    def __init__(self, project_name, project_category, project_type, type_code, project_zip, customer, quote, terms, tax, billing, labor_code, order_type, price):
        self.project_number = get_next_project_num()
        self.project_name = project_name
        self.project_category = project_category
        self.project_type = project_type
        self.type_code = type_code
        self.project_zip = project_zip
        self.customer = customer
        self.quote = quote
        self.terms = terms
        self.tax = tax  # set as  boolean
        self.billing = billing
        self.labor_code = labor_code
        self.order_type = order_type
        self.price = price

        project = f'{self.project_number} {self.project_name} {self.type_code}'

        self.write_log()

        # make new year folder if current year does not match dir list-------------
        project_dir_list = (os.listdir(project_dir))
        year_list = []

        for year in project_dir_list:
            try:
                if int(year[:5]) and len(year) < 5:
                    year_list.append(year)
            except ValueError:
                continue
        year_list.sort()
        if year_list[-1] == str(current_year):
            year_dir = current_year
        else:
            self.createFolder(f'{project_dir}/{current_year}')
        #--------------------------------------------------------------------------

        self.createFolder(f'{project_dir}/{current_year}/{project}/photos')
        self.createFolder(f'{project_dir}/{current_year}/{project}/test_reports')
        self.createFolder(f'{project_dir}/{current_year}/{project}/insurance_docs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/minutes_etc')
        self.createFolder(f'{project_dir}/{current_year}/{project}/travel')
        self.createFolder(f'{project_dir}/{current_year}/{project}/tx')
        self.createFolder(f'{project_dir}/{current_year}/{project}/billing')
        self.createFolder(f'{project_dir}/{current_year}/{project}/billing/QB_Invoices')
        self.createFolder(f'{project_dir}/{current_year}/{project}/production')
        self.createFolder(f'{project_dir}/{current_year}/{project}/RFIs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/Purchasing')
        self.createFolder(f'{project_dir}/{current_year}/{project}/Material_Specs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/quotes')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/change_orders')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/closeout_documents')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/AIA_docs_for_pay_apps')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/backup_and_old_files')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/Exhibits')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/TX_Ex')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/drawings_sent')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/revisions')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/archive_dwgs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/approved_dwgs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/arch_dwgs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/solidworks')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/typical_drawings')
        self.createFolder(f'{project_dir}/{current_year}/{project}/safety')

        if windows_user != 'mark':
            self.make_readme(project)
        
        self.path =  f"{project_dir}\{current_year}\{project}"

        start = time.strftime("%#m/%#d/%y") # current date without leading zeros
        task_list = [['Submittal Drawings', 'Bryan Bordeman', (datetime.strptime(start, '%m/%d/%y').date() + timedelta(days=7)).strftime("%#m/%#d/%y")],
                     ['Review Drawings', 'Mark Holder', (datetime.strptime(
                         start, '%m/%d/%y').date() + timedelta(days=5)).strftime("%#m/%#d/%y")],
                     ['Review Contract', 'Mark Holder', start],
                     ['Quickbooks – Create New Estimate', 'Mark Holder', start],
                     ['Invoice', 'Mark Holder', start],
                     ['Insurance documents', 'Jori Goldner', start],
                     ['Project schedule', 'Bryan Bordeman', start]]

        self.create_task(self.project_number, task_list)

    def write_project_data(self):
        '''write opportunity to database'''

        table = 'project'

        self.execute_db_query(f"""INSERT INTO {table} (
                                    project_number,
                                    project_name,
                                    project_category,
                                    project_type,
                                    type_code,
                                    project_zip,
                                    customer,
                                    quote,
                                    terms,
                                    tax,
                                    billing,
                                    labor_code,
                                    order_type,
                                    price
                                    ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);""", (str(self.project_number),
                                    str(self.project_name),
                                    str(self.project_category),
                                    str(self.project_type),
                                    str(self.type_code),
                                    str(self.project_zip),
                                    str(self.customer),
                                    str(self.quote),
                                    str(self.terms),
                                    str(self.tax),
                                    str(self.billing),
                                    str(self.labor_code),
                                    str(self.order_type),
                                    str(self.price)))

    def __repr__(self):
        return f'{self.project_number} {self.project_name} {self.type_code}'

    def write_log(self):

        log = "Global Job List-start-complete dates 2016"

        completeName = os.path.join(project_dir, log + ".xlsx")


        book = load_workbook(completeName)
        ws = book.worksheets[0]
        for cell in ws["B"]:
            if cell.value is None:
                current_row = cell.row
                break
        else:
            cell.row + 1

        ws[f'A{current_row}'] = f'{self.project_number}{self.billing}'
        ws[f'B{current_row}'] = find_state(int(self.project_zip)).split(', ')[-1]
        ws[f'C{current_row}'] = self.order_type
        ws[f'D{current_row}'] = f'{self.project_name} {self.type_code}'
        ws[f'E{current_row}'] = self.labor_code
        ws[f'F{current_row}'] = time.strftime("%D")
        ws[f'G{current_row}'] = self.project_type

        book.save(completeName)
        book.close()
        self.create_table()
        self.write_project_data()



    def createFolder(self, directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print('Error: Creating directory. ' + directory)

    def make_readme(self, project):
        name_of_file = 'README'
        path = f'{project_dir}/{current_year}/{project}'
        completeName = os.path.join(path, name_of_file + ".txt")

        readme = open(completeName, "w")

        # convert boolean to yes/ no string for readme
        if self.tax == 1:
            self.tax = 'Yes'
        elif self.tax == 0:
            self.tax = 'No'

        # convert billing back to key for readme
        for key, value in billing_dict.items():
            if value == self.billing:
                self.billing = key

        project_info = f'Project Number = {self.project_number}\nProject Name = {self.project_name}\nProject Category = {self.project_category}\nProject Type = {self.project_type}\nProject Zip = {self.project_zip}\nCustomer = {self.customer}\nQuote Number = {self.quote}\nTerms = {self.terms}\nTax Exempt = {self.tax}\nBilling Type = {self.billing}\nSell Price (USD) = ${self.price}\n'
        readme.write(project_info)
        readme.close()

class Service(Opportunity):
    def __init__(self, project_name, project_category, project_type, type_code, project_zip, customer, quote, terms, tax, billing, labor_code, order_type, price):
        self.project_number = get_next_service_num()
        self.project_name = project_name
        self.project_category = project_category  # readonly on GUI (set Door Service)
        self.project_type = project_type  # because project_category is set limited to choices in category
        self.type_code = type_code
        self.project_zip = project_zip
        self.customer = customer
        self.quote = quote
        self.terms = terms
        self.tax = tax  # set as  boolean
        self.billing = billing
        self.labor_code = labor_code
        self.order_type = order_type
        self.price = price

        project = f'{self.project_number} {self.project_name} {self.type_code}'


        # make new year folder if current year does not match dir list-------------
        project_dir_list = (os.listdir(rf'{project_dir}/Door Service'))
        year_list = []

        for year in project_dir_list:
            if year[:5].isdigit() and len(year) < 5:
                    year_list.append(year)
        year_list.sort()
        if year_list[-1] == str(current_year):
            year_dir = current_year
        else:
            self.createFolder(f'{project_dir}/Door Service/{current_year}')
        #--------------------------------------------------------------------------

        self.createFolder(f'{project_dir}/Door Service/{current_year}/{project}/photos')
        self.createFolder(f'{project_dir}/Door Service/{current_year}/{project}/billing')
        self.createFolder(f'{project_dir}/Door Service/{current_year}/{project}/billing/QB_Invoices')
        self.createFolder(f'{project_dir}/Door Service/{current_year}/{project}/billing/payments')
        self.createFolder(f'{project_dir}/Door Service/{current_year}/{project}/production')
        self.createFolder(f'{project_dir}/Door Service/{current_year}/{project}/contracts')
        self.createFolder(f'{project_dir}/Door Service/{current_year}/{project}/reports')


        self.make_readme(project)

        start = time.strftime("%#m/%#d/%y") # current date without leading zeros
        task_list = [['Quickbooks – Create New Estimate', 'Mark Holder', start],
                     ['Invoice', 'Mark Holder', start],
                     ['Project schedule', 'Bryan Bordeman', start]]

        self.create_task(f'SVC{self.project_number}', task_list)

        self.write_project_data()

    def write_project_data(self):
        '''write opportunity to database'''

        table = 'project'

        self.execute_db_query(f"""INSERT INTO {table} (
                                    project_number,
                                    project_name,
                                    project_category,
                                    project_type,
                                    type_code,
                                    project_zip,
                                    customer,
                                    quote,
                                    terms,
                                    tax,
                                    billing,
                                    labor_code,
                                    order_type,
                                    price
                                    ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);""", (str(self.project_number),
                                    str(self.project_name),
                                    str(self.project_category),
                                    str(self.project_type),
                                    str(self.type_code),
                                    str(self.project_zip),
                                    str(self.customer),
                                    str(self.quote),
                                    str(self.terms),
                                    str(self.tax),
                                    str(self.billing),
                                    str(self.labor_code),
                                    str(self.order_type),
                                    str(self.price)))

    def __repr__(self):
        return f'{self.project_number} {self.project_name} {self.type_code}'

    def createFolder(self, directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print('Error: Creating directory. ' + directory)

    def make_readme(self, project):
        name_of_file = 'README'
        path = f'{project_dir}/Door Service/{current_year}/{project}'
        completeName = os.path.join(path, name_of_file + ".txt")

        readme = open(completeName, "w")

        # convert boolean to yes/ no string for readme
        if self.tax == 1:
            self.tax = 'Yes'
        elif self.tax == 0:
            self.tax = 'No'

        # convert billing back to key for readme
        for key, value in billing_dict.items():
            if value == self.billing:
                self.billing = key

        project_info = f'Project Number = {self.project_number}\nProject Name = {self.project_name}\nProject Category = {self.project_category}\nProject Type = {self.project_type}\nProject Zip = {self.project_zip}\nCustomer = {self.customer}\nQuote Number = {self.quote}\nTerms = {self.terms}\nTax Exempt = {self.tax}\nBilling Type = {self.billing}\nSell Price (USD) = ${self.price}\n'
        readme.write(project_info)
        readme.close()


if __name__ == "__main__":
    main()
