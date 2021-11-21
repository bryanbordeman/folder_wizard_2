# import sys
# from create_opportunity import create_opportunity
# from zip_2_state import find_state
from openpyxl import*
import os
import os.path

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
log = os.path.join(BASE_DIR, "test.xlsx")

# def main():
#     print(create_log())
#     sys.stdout.flush()

# def create_log():
#     quote = create_opportunity() # make Opportunty object
#     # write Excel log 
#     # state = find_state(int(quote.project_zip))
    

#     return quote

def main():
    create_log()

# def create_log():
#             '''update quote log'''
#         log = "Master Quote Log"
#         completeName = os.path.join(opportunity_dir, log + ".xlsx")
#         book = load_workbook(completeName)
#         ws = book.worksheets[0]
#         for cell in ws["A"]:
#             if cell.value is None:
#                 print(cell.row)
#                 current_row = cell.row
#                 break
#         else:
#             cell.row + 1

#         ws[f'A{current_row}'] = quote_number
#         ws[f'B{current_row}'] = manager
#         ws[f'C{current_row}'] = time.strftime("%D")
#         ws[f'D{current_row}'] = f'{self.project_name} {self.type_code}'
#         ws[f'E{current_row}'] = find_state(int(self.project_zip))
#         ws[f'F{current_row}'] = self.bid_due

#         #NOTE below is code to fill cell with green background

#         # for col in ['A','B','C','D','E','F','G','H']:
#         #     ws[f'{col}{current_row}'].fill = PatternFill(
#         #         start_color="92d050", end_color="92d050", fill_type="solid")

#         book.save(completeName)
#         book.close()
    
def create_log():
    book = load_workbook(log) #load file
    ws = book.worksheets[0] #load worksheet
    ws[f'A{1}'] = "quote_number"
    book.save(log)
    book.close()

if __name__ == "__main__":
    main()