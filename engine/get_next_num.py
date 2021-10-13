'''==========================================
Title:  get_next_num.py
Author:  Bryan Bordeman
Start Date:  062219
Updated:  020220 (added next_service_num())
Version:  support script

;=========================================='''


from openpyxl import*
import time
import os

# below path is for testing at home
# project_dir = r'C:\Users\Bryan\Google Drive\Programming\Python\folder_wizard\Global'
# opportunity_dir = r"C:\Users\Bryan\Google Drive\Programming\Python\folder_wizard\RFQ's"

# below path is live on office server
project_dir = r'T:\Global'
opportunity_dir = r"T:\RFQ's"

# global variables
current_year = time.strftime("%Y")


def main():
    print(get_next_quote_num())
    print(get_next_project_num())
    print(get_next_service_num())



def get_next_quote_num():
    current_quote = int()
    current_quote_year = ''
    try:
        log = "Master Quote Log"

        completeName = os.path.join(opportunity_dir, log + ".xlsx")

        book = load_workbook(completeName)
        ws = book.worksheets[0]

        year = time.strftime("%Y")[2:]

        # below is how next next number is determined

        for cell in ws["A"]:
            if cell.value is not None and cell.value != "Quote" and cell.value[1:3] == year:
                    if int((cell.value)[4:]) > current_quote:
                        current_quote = int((cell.value)[4:])
                        current_quote_year = str((cell.value)[1:3])

        #--------------------------------------------------------------------
        #NOTE OLD CODE NOT USED (was taking number based on last empty cell)
        # for cell in ws["A"]:
        #     if cell.value is None:
        #         current_quote = int((ws[f"A{cell.row - 1}"].value)[4:])
        #         current_quote_year = (ws[f"A{cell.row - 1}"].value)[1:3]
        #         break
        # else:
        #     cell.row + 1

        #--------------------------------------------------------------------

        if current_quote_year == year:
            next_quote = str(current_quote + 1)
            for i in range(3):
                if len(next_quote) < 3:
                    next_quote = '0' + next_quote
            number = f'Q{year}-{next_quote}'
            next_quote = current_quote + 1
        else:
            next_quote = '001'
            number = f'Q{year}-{next_quote}'
        return number

        book.close()

    except FileNotFoundError:
        return None

    # finally:
    #     book.close()


def get_next_project_num():
    current_project = ''
    try:
        log = "Global Job List-start-complete dates 2016"

        completeName = os.path.join(project_dir, log + ".xlsx")

        book = load_workbook(completeName)

        ws = book.worksheets[0]
        for cell in ws["B"]:
            if cell.value is None:
                current_project = int((ws[f"A{cell.row}"].value))
                break
        else:
            cell.row + 1
        return current_project

        book.close()

    except FileNotFoundError:
        return None

    # finally:
    #     book.close()

def get_next_service_num():


    try:
        project_dir_list = (os.listdir(rf'{project_dir}/Door Service/{current_year}'))

    except FileNotFoundError:
        createFolder(f'{project_dir}/Door Service/{current_year}')
        project_dir_list = (os.listdir(rf'{project_dir}/Door Service/{current_year}'))

    service_number_list = []

    for folder in project_dir_list:
        if folder[:5].isdigit():
            service_number_list.append(int(folder[:5]))

    if len(service_number_list) > 0:
        next_service_num = service_number_list[-1] + 1
    else:
        # if no folders are in dir
        next_service_num = str(current_year[-2:]) + '001'

    return str(next_service_num)


def createFolder(directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print('Error: Creating directory. ' + directory)

if __name__ == "__main__":
    main()
